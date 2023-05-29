import fastwer
import openpyxl

ExcelName =  "Results.xlsx"

# Loading the results 
readingDataframe = openpyxl.load_workbook(ExcelName, data_only=True)
readingSheet = readingDataframe.worksheets[3]         #The OCR results are on the fourth sheet

writingDataframe = openpyxl.load_workbook(ExcelName)
writingSheet = writingDataframe.worksheets[3]  

for i in range(2,851):
    ocrText = readingSheet.cell(row = i, column = 8).value
    groundTruth = readingSheet.cell(row = i, column = 9).value

    if (ocrText == None):
        ocrText = ""
    if (groundTruth ==  None):
        groundTruth = ""
    
    ocrText = ocrText.lower()
    groundTruth = groundTruth.lower()

    cre = fastwer.score_sent(ocrText, groundTruth, char_level=True)
    writingSheet.cell(row=i, column=10).value = cre

writingDataframe.save(ExcelName)
writingDataframe.close()